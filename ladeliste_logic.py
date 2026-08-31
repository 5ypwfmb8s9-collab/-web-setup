"""Verarbeitungslogik fuer das Ladelisten-Tool.

Enthaelt die komplette, UI-unabhaengige Logik:
- Zusammenfuehren mehrerer Quelldateien (Blatt "Avis Lade Listen")
- Aufbau der LKW-Blaetter nach der verbindlichen Spezifikation
- Validierung des Ergebnisses

Kann sowohl von der Streamlit-Oberflaeche (app.py) als auch direkt per
Kommandozeile aufgerufen werden:

    python ladeliste_logic.py quelle1.xlsx quelle2.xlsx -o ausgabe.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, OrderedDict
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Konstanten
# ---------------------------------------------------------------------------

SOURCE_SHEET_NAME = "Avis Lade Listen"

# 0-basierte Spaltenindizes in den ROHEN Quelldateien (A=0, B=1, ...), wie sie
# hochgeladen werden. "Plan VL" (Spalte K) ist die LKW-Kennung - Spalte L
# "LKW VL" ist in der Praxis oft leer und darf NICHT verwendet werden.
RAW_COL_ABHOLTAG = 3      # D
RAW_COL_EMPF_NAME = 6     # G
RAW_COL_EMPF_ORT = 7      # H
RAW_COL_ABLADESTELLE = 8  # I
RAW_COL_PLAN_VL = 10      # K ("Plan VL")
RAW_COL_PAL_QUELLE = 13   # N
RAW_COL_BRUTTO = 14       # O

# Feste 7 Kernspalten, sowohl im zusammengefuehrten "Avis Lade Listen"-Blatt
# als auch in jedem LKW-Blatt (0-basierte Indizes innerhalb dieser Spaltenliste).
TRUCK_SHEET_HEADERS = [
    "Abholtag",
    "Empfaenger Name",
    "Empfaenger Ort",
    "Abladestelle",
    "Plan VL",
    "PAL",
    "Brutto",
]

COL_ABHOLTAG = 0
COL_EMPF_NAME = 1
COL_EMPF_ORT = 2
COL_ABLADESTELLE = 3
COL_PLAN_VL = 4
COL_PAL_QUELLE = 5
COL_BRUTTO = 6

PAL_TOKENS = ("VWPAL", "111444")

PAL_PATTERN = re.compile(r"(\d+)\s*\*\s*(VWPAL|111444)(?![A-Za-z0-9])", re.IGNORECASE)

# Feste Spaltenbreiten je Blatt (in "pt", wie in der Spezifikation angegeben)
COLUMN_WIDTHS_PT = {
    "A": 85,
    "B": 210,
    "C": 105,
    "D": 100,
    "E": 85,
    "F": 180,
    "G": 85,
}

HEADER_FILL_COLOR = "FF1F4E78"
HEADER_FONT_COLOR = "FFFFFFFF"
HEADER_BORDER_COLOR = "FF404040"
DATA_BORDER_COLOR = "FFD9D9D9"

DATE_NUMBER_FORMAT = "DD.MM.YYYY"
# [$-407] erzwingt die deutsche Anzeige (Punkt=Tausender, Komma=Dezimal)
# unabhaengig von der Locale des oeffnenden Excel/LibreOffice.
WEIGHT_NUMBER_FORMAT = "[$-407]#,##0.0"
QUANTITY_NUMBER_FORMAT = "0"

CONFIRMATION_TEXTS = [
    "Übernahmebestätigung des Fahrers/Warenempfängers.",
    "Der Fahrer des LKW's bestätigt eine ordnungsgemäße Ladungssicherung durchgeführt zu haben",
    "und bestätigt außerdem, das zulässige Ladungsgewicht nicht überschritten zu haben.",
]

MAX_SHEET_NAME_LENGTH = 31

# ---------------------------------------------------------------------------
# Planungsabgleich (Planung fuer Staplerfahrer)
# ---------------------------------------------------------------------------

PLANUNG_SHEET_NAME = "Planung für Staplerfahrer"
ABWEICHUNGEN_SHEET_NAME = "Abweichungen"

# 1-basierte Spaltenindizes im Blatt "Planung fuer Staplerfahrer".
PLANUNG_COL_KAPI = 2           # B - Abladestelle
PLANUNG_COL_PALET_SAYISI = 8   # H - Menge
PLANUNG_COL_PALET = 9          # I - Ladungstraeger-Art (VWPAL/111444/...)
PLANUNG_COL_LKW = 10           # J - LKW-Kennzeichen (wird neu befuellt)
PLANUNG_HEADER_ROW = 4
PLANUNG_DATA_START_ROW = 5

# Abladestellen, die grundsaetzlich nie in den Avis-Listen vorkommen und
# daher beim Abgleich (Befuellen der LKW-Spalte und Abweichungspruefung)
# uebersprungen werden.
IGNORED_ABLADESTELLEN = {"KXG", "BAU90", "BAU50", "CTCT", "UPSOR", "ACHG", "PT02"}


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def cm_to_inch(cm: float) -> float:
    return cm / 2.54


def pt_to_excel_width(pt: float) -> float:
    """Rechnet eine Punktangabe naeherungsweise in Excel-Spaltenbreiten-Einheiten um.

    Excel misst Spaltenbreiten nicht in Punkt, sondern in "Zeichenbreiten" der
    Standardschrift. Diese Naeherung (ueber Pixel, 96dpi, ~7px Zeichenbreite)
    liefert fuer alle Blaetter deterministisch identische Werte.
    """
    pixels = pt * 96 / 72
    width = (pixels - 5) / 7
    return round(width, 2)


def _clean_str(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_lkw(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def to_number(value: Any) -> Optional[float]:
    """Wandelt einen Quellwert in eine echte Zahl (int/float) um."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return None
    normalized = text.replace(".", "").replace(",", ".") if "," in text else text
    try:
        num = float(normalized)
    except ValueError:
        try:
            num = float(text)
        except ValueError:
            return None
    if num.is_integer():
        return int(num)
    return num


def to_date(value: Any):
    """Gibt ein echtes date/datetime-Objekt zurueck, falls moeglich."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return value


def process_pal(text: Any) -> str:
    """Extrahiert VWPAL/111444-Mengen aus einem Ladungstraeger-Text.

    Nur die Ladungstraeger "VWPAL" und "111444" werden beruecksichtigt, alle
    anderen Angaben werden entfernt. Die Erkennung ist robust gegenueber
    beliebigen Leerzeichen um "*" und "/". Reihenfolge im Ergebnis = Reihenfolge
    des ersten Vorkommens der jeweiligen Art im Quelltext.
    """
    if text is None:
        return ""
    source = str(text)
    matches = list(PAL_PATTERN.finditer(source))
    if not matches:
        return ""

    totals: Dict[str, int] = {}
    first_pos: Dict[str, int] = {}
    for m in matches:
        qty = int(m.group(1))
        token = m.group(2).upper()
        totals[token] = totals.get(token, 0) + qty
        if token not in first_pos:
            first_pos[token] = m.start()

    ordered = sorted(totals.keys(), key=lambda t: first_pos[t])
    return " / ".join(f"{totals[t]}*{t}" for t in ordered)


# ---------------------------------------------------------------------------
# Schritt 1: Einlesen & Zusammenfuehren
# ---------------------------------------------------------------------------

def _get(row: Sequence[Any], idx: int) -> Any:
    return row[idx] if idx < len(row) else None


def read_source_rows(file: Any, sheet_name: str = SOURCE_SHEET_NAME) -> List[List[Any]]:
    """Liest Datenzeilen aus dem angegebenen Blatt einer Rohdatei und bildet sie
    direkt auf die 7 Kernspalten ab (Abholtag, Empfaenger Name, Empfaenger Ort,
    Abladestelle, Plan VL, PAL, Brutto). Die PAL-Spalte bleibt dabei unbearbeitet
    (roher Text aus Spalte N).

    `file` kann ein Pfad, ein Datei-Objekt oder ein BytesIO sein (alles, was
    openpyxl.load_workbook akzeptiert).
    """
    wb = load_workbook(file, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Das Blatt '{sheet_name}' wurde in der Datei nicht gefunden."
        )
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        next(rows_iter)  # Kopfzeile der Rohdatei ueberspringen
    except StopIteration:
        wb.close()
        return []

    data: List[List[Any]] = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        mapped = [
            _get(row, RAW_COL_ABHOLTAG),
            _get(row, RAW_COL_EMPF_NAME),
            _get(row, RAW_COL_EMPF_ORT),
            _get(row, RAW_COL_ABLADESTELLE),
            _get(row, RAW_COL_PLAN_VL),
            _get(row, RAW_COL_PAL_QUELLE),
            _get(row, RAW_COL_BRUTTO),
        ]
        data.append(mapped)
    wb.close()
    return data


def merge_source_files(files: Sequence[Any]) -> Tuple[List[Any], List[List[Any]]]:
    """Fuehrt die Datenzeilen mehrerer Quelldateien zusammen (bereits auf die 7
    Kernspalten abgebildet).

    Reihenfolge: alle Zeilen der ersten Datei zuerst (Originalreihenfolge),
    danach die der naechsten usw. Keine Sortierung/Deduplizierung/Gruppierung.
    """
    if not files:
        raise ValueError("Es wurde keine Quelldatei uebergeben.")

    combined_rows: List[List[Any]] = []
    for f in files:
        combined_rows.extend(read_source_rows(f))

    return list(TRUCK_SHEET_HEADERS), combined_rows


def write_source_sheet(wb: Workbook, header: Sequence[Any], rows: Sequence[Sequence[Any]]) -> Worksheet:
    """Schreibt das zusammengefuehrte Ergebnis unveraendert als Quellblatt."""
    ws = wb.create_sheet(SOURCE_SHEET_NAME)
    if header:
        ws.append(list(header))
    for row in rows:
        ws.append(list(row))
    return ws


# ---------------------------------------------------------------------------
# Schritt 2: LKW-Blaetter
# ---------------------------------------------------------------------------

@dataclass
class TruckEntry:
    abholtag: Any
    empf_name: Any
    empf_ort: Any
    abladestelle: Any
    plan_vl: str
    pal: str
    brutto: Optional[float]
    pal_raw: Any


@dataclass
class TruckSheetSummary:
    sheet_name: str
    plan_vl: str
    position_count: int
    gesamtgewicht: float
    total_111444: int
    total_vwpal: int


def build_truck_groups(rows: Sequence[Sequence[Any]]) -> "OrderedDict[str, List[TruckEntry]]":
    """Gruppiert Zeilen nach Spalte "Plan VL", Zeilen mit leerem Plan VL werden
    ignoriert. Der Wert wird unveraendert (inkl. eines eventuellen "-N"-Zusatzes
    wie "BOHDT978-2") als eigenstaendige Kennung verwendet - kein Abschneiden
    von Suffixen, jede Auspraegung bekommt ihr eigenes Blatt. Reihenfolge der
    Zeilen je Gruppe = Reihenfolge im Quellblatt. Reihenfolge der Gruppen =
    Reihenfolge des ersten Auftretens im Quellblatt.
    """
    groups: "OrderedDict[str, List[TruckEntry]]" = OrderedDict()
    for row in rows:
        plan_vl = normalize_lkw(_get(row, COL_PLAN_VL))
        if plan_vl is None:
            continue
        entry = TruckEntry(
            abholtag=to_date(_get(row, COL_ABHOLTAG)),
            empf_name=_get(row, COL_EMPF_NAME),
            empf_ort=_get(row, COL_EMPF_ORT),
            abladestelle=_get(row, COL_ABLADESTELLE),
            plan_vl=plan_vl,
            pal=process_pal(_get(row, COL_PAL_QUELLE)),
            brutto=to_number(_get(row, COL_BRUTTO)),
            pal_raw=_get(row, COL_PAL_QUELLE),
        )
        groups.setdefault(plan_vl, []).append(entry)
    return groups


def _truck_sheet_name(plan_vl: str) -> str:
    name = f"LKW {plan_vl}"
    if len(name) > MAX_SHEET_NAME_LENGTH:
        name = name[:MAX_SHEET_NAME_LENGTH]
    return name


# Ausgeblendete Helferspalten (H, I) je LKW-Blatt: pro Datenzeile wird dort
# die Menge 111444 bzw. VWPAL aus der jeweiligen F-Zelle extrahiert. Damit
# bestehen die Summenformeln nur noch aus einem einfachen SUM() ueber diese
# Spalten - die textauswertende Formel bezieht sich dabei je Zeile nur auf
# eine einzelne Zelle (kein Bereich/Array-Kontext), was sie unabhaengig von
# der jeweiligen Excel-Version garantiert fehlerfrei auswertbar macht.
HELPER_COL_111444 = 8  # H
HELPER_COL_VWPAL = 9   # I


def _pal_qty_cell_formula(token: str, cell_ref: str) -> str:
    """Baut eine einfache Formel, die die Menge einer PAL-Art (VWPAL oder
    111444) aus EINER einzelnen Zelle extrahiert - unabhaengig davon, ob die
    Art als erster oder zweiter Eintrag im Zelltext auftritt. Da sich die
    Formel nur auf eine einzelne Zelle bezieht, ist sie ohne Array-Formel-
    Eigenheiten (SUMPRODUCT-Trick) ueberall identisch auswertbar.
    """
    marker = f"*{token}"
    prefix = f'LEFT({cell_ref},FIND("{marker}",{cell_ref})-1)'
    has_slash = f'ISNUMBER(SEARCH("/",{prefix}))'
    after_slash = f'TRIM(MID({prefix},SEARCH("/",{prefix})+1,100))'
    no_slash = f"TRIM({prefix})"
    value_expr = f"VALUE(IF({has_slash},{after_slash},{no_slash}))"
    return f"=IFERROR({value_expr},0)"


def _style_header_row(ws: Worksheet) -> None:
    thin_dark = Side(style="thin", color=HEADER_BORDER_COLOR)
    border = Border(left=thin_dark, right=thin_dark, top=thin_dark, bottom=thin_dark)
    font = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_COLOR)
    fill = PatternFill(fill_type="solid", start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, title in enumerate(TRUCK_SHEET_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border

    ws.row_dimensions[1].height = 24


DATA_COL_ALIGNMENT = {
    "A": Alignment(horizontal="center", vertical="center", wrap_text=False),
    "B": Alignment(horizontal="left", vertical="center", wrap_text=True),
    "C": Alignment(horizontal="left", vertical="center", wrap_text=True),
    "D": Alignment(horizontal="left", vertical="center", wrap_text=True),
    "E": Alignment(horizontal="center", vertical="center", wrap_text=False),
    "F": Alignment(horizontal="left", vertical="center", wrap_text=True),
    "G": Alignment(horizontal="right", vertical="center", wrap_text=False),
}


def _style_data_rows(ws: Worksheet, first_row: int, last_row: int) -> None:
    thin_light = Side(style="thin", color=DATA_BORDER_COLOR)
    border = Border(left=thin_light, right=thin_light, top=thin_light, bottom=thin_light)
    font = Font(name="Calibri", size=10, bold=False, color="FF000000")

    for row in range(first_row, last_row + 1):
        for col_letter in ("A", "B", "C", "D", "E", "F", "G"):
            cell = ws[f"{col_letter}{row}"]
            cell.font = font
            cell.border = border
            cell.alignment = DATA_COL_ALIGNMENT[col_letter]


def _write_data_row(ws: Worksheet, row_idx: int, entry: TruckEntry) -> None:
    ws.cell(row=row_idx, column=1, value=entry.abholtag)
    ws.cell(row=row_idx, column=2, value=entry.empf_name)
    ws.cell(row=row_idx, column=3, value=entry.empf_ort)
    ws.cell(row=row_idx, column=4, value=entry.abladestelle)
    ws.cell(row=row_idx, column=5, value=entry.plan_vl)
    ws.cell(row=row_idx, column=6, value=entry.pal if entry.pal else None)
    ws.cell(row=row_idx, column=7, value=entry.brutto)

    ws.cell(row=row_idx, column=1).number_format = DATE_NUMBER_FORMAT
    ws.cell(row=row_idx, column=7).number_format = WEIGHT_NUMBER_FORMAT


def build_truck_sheet(wb: Workbook, plan_vl: str, entries: List[TruckEntry]) -> Tuple[Worksheet, TruckSheetSummary]:
    sheet_name = _truck_sheet_name(plan_vl)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    _style_header_row(ws)

    first_data_row = 2
    last_data_row = first_data_row + len(entries) - 1
    for offset, entry in enumerate(entries):
        row_idx = first_data_row + offset
        _write_data_row(ws, row_idx, entry)
        ws.cell(row=row_idx, column=HELPER_COL_111444,
                value=_pal_qty_cell_formula("111444", f"F{row_idx}"))
        ws.cell(row=row_idx, column=HELPER_COL_VWPAL,
                value=_pal_qty_cell_formula("VWPAL", f"F{row_idx}"))

    ws.column_dimensions[get_column_letter(HELPER_COL_111444)].hidden = True
    ws.column_dimensions[get_column_letter(HELPER_COL_VWPAL)].hidden = True

    _style_data_rows(ws, first_data_row, last_data_row)

    n = last_data_row
    row_gesamtgewicht = n + 1
    row_111444 = n + 2
    row_vwpal = n + 3
    row_gesamt = n + 4
    # n + 5 bleibt eine volle Leerzeile vor dem Bestaetigungsbereich.
    row_conf = [n + 6, n + 7, n + 8]
    row_signature = n + 12

    bold_font = Font(name="Calibri", size=10, bold=True, color="FF000000")
    regular_font = Font(name="Calibri", size=10, bold=False, color="FF000000")
    right_align = Alignment(horizontal="right", vertical="center")

    # Gesamtgewicht
    c_label = ws.cell(row=row_gesamtgewicht, column=6, value="Gesamtgewicht:")
    c_label.font = bold_font
    c_label.alignment = right_align
    c_value = ws.cell(row=row_gesamtgewicht, column=7, value=f"=SUM(G{first_data_row}:G{n})")
    c_value.font = bold_font
    c_value.alignment = right_align
    c_value.number_format = WEIGHT_NUMBER_FORMAT

    # 111444
    col_111444 = get_column_letter(HELPER_COL_111444)
    col_vwpal = get_column_letter(HELPER_COL_VWPAL)
    e_label = ws.cell(row=row_111444, column=5, value="111444")
    e_label.font = regular_font
    f_value = ws.cell(row=row_111444, column=6, value=f"=SUM({col_111444}{first_data_row}:{col_111444}{n})")
    f_value.font = regular_font
    f_value.alignment = right_align
    f_value.number_format = QUANTITY_NUMBER_FORMAT

    # VWPAL
    e_label2 = ws.cell(row=row_vwpal, column=5, value="VWPAL")
    e_label2.font = regular_font
    f_value2 = ws.cell(row=row_vwpal, column=6, value=f"=SUM({col_vwpal}{first_data_row}:{col_vwpal}{n})")
    f_value2.font = regular_font
    f_value2.alignment = right_align
    f_value2.number_format = QUANTITY_NUMBER_FORMAT

    # Gesamt
    e_label3 = ws.cell(row=row_gesamt, column=5, value="Gesamt:")
    e_label3.font = bold_font
    f_value3 = ws.cell(row=row_gesamt, column=6, value=f"=F{row_111444}+F{row_vwpal}")
    f_value3.font = bold_font
    f_value3.alignment = right_align
    f_value3.number_format = QUANTITY_NUMBER_FORMAT

    top_border = Border(top=Side(style="thin", color="FF000000"))
    e_label3.border = top_border
    f_value3.border = top_border

    # Bestaetigungsbereich
    conf_font = Font(name="Calibri", size=10, bold=False, color="FF000000")
    conf_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r, text in zip(row_conf, CONFIRMATION_TEXTS):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = conf_font
        cell.alignment = conf_alignment

    # Unterschriftsbereich (drei Leerzeilen davor, keine Verbindung)
    sig_font = Font(name="Calibri", size=10, bold=False, color="FF000000")
    sig_left = Alignment(horizontal="left", vertical="bottom")
    sig_right = Alignment(horizontal="right", vertical="bottom")

    a_sig = ws.cell(row=row_signature, column=1, value="Unterschrift Staplerfahrer")
    a_sig.font = sig_font
    a_sig.alignment = sig_left

    d_sig = ws.cell(row=row_signature, column=4, value="Unterschrift Fahrer")
    d_sig.font = sig_font
    d_sig.alignment = sig_left

    g_sig = ws.cell(row=row_signature, column=7, value="Kennzeichen LKW")
    g_sig.font = sig_font
    g_sig.alignment = sig_right

    # Spaltenbreiten (kein AutoFit)
    for col_letter, pt in COLUMN_WIDTHS_PT.items():
        ws.column_dimensions[col_letter].width = pt_to_excel_width(pt)

    _apply_print_settings(ws, row_signature)

    summary = TruckSheetSummary(
        sheet_name=sheet_name,
        plan_vl=plan_vl,
        position_count=len(entries),
        gesamtgewicht=sum((e.brutto or 0) for e in entries),
        total_111444=sum(_extract_qty(e.pal, "111444") for e in entries),
        total_vwpal=sum(_extract_qty(e.pal, "VWPAL") for e in entries),
    )
    return ws, summary


def _extract_qty(pal_text: str, token: str) -> int:
    if not pal_text:
        return 0
    m = re.search(rf"(\d+)\*{token}", pal_text)
    return int(m.group(1)) if m else 0


def _apply_print_settings(ws: Worksheet, last_row: int) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.print_options.gridLines = False
    ws.print_options.headings = False

    ws.print_area = f"A1:G{last_row}"
    ws.print_title_rows = "1:1"

    ws.page_margins.left = cm_to_inch(0.64)
    ws.page_margins.right = cm_to_inch(0.64)
    ws.page_margins.top = cm_to_inch(1.91)
    ws.page_margins.bottom = cm_to_inch(1.91)
    ws.page_margins.header = cm_to_inch(0.76)
    ws.page_margins.footer = cm_to_inch(0.76)


# ---------------------------------------------------------------------------
# Schritt 3: Validierung
# ---------------------------------------------------------------------------

@dataclass
class ValidationResult:
    check: str
    passed: bool
    message: str


ERROR_VALUES = {"#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}


def validate(
    header: Sequence[Any],
    rows: Sequence[Sequence[Any]],
    groups: "OrderedDict[str, List[TruckEntry]]",
    wb: Workbook,
) -> List[ValidationResult]:
    results: List[ValidationResult] = []

    # 1 + 4: jede Zeile mit gefuelltem Plan VL ist genau einem LKW-Blatt
    # zugeordnet, Positionsanzahl je LKW stimmt.
    expected_counts: Dict[str, int] = OrderedDict()
    total_considered = 0
    for row in rows:
        plan_vl = normalize_lkw(_get(row, COL_PLAN_VL))
        if plan_vl is None:
            continue
        total_considered += 1
        expected_counts[plan_vl] = expected_counts.get(plan_vl, 0) + 1

    assigned_total = sum(len(v) for v in groups.values())
    results.append(ValidationResult(
        "Zeilenzuordnung vollstaendig",
        assigned_total == total_considered,
        f"{assigned_total} von {total_considered} Zeilen mit Plan VL zugeordnet.",
    ))

    count_mismatches = [
        plan_vl for plan_vl, count in expected_counts.items()
        if count != len(groups.get(plan_vl, []))
    ]
    results.append(ValidationResult(
        "Positionsanzahl je LKW korrekt",
        not count_mismatches,
        "OK" if not count_mismatches else f"Abweichung bei: {', '.join(count_mismatches)}",
    ))

    # 2: kein LKW-Blatt doppelt
    truck_sheet_names = [_truck_sheet_name(plan_vl) for plan_vl in groups.keys()]
    duplicates = [n for n in set(truck_sheet_names) if truck_sheet_names.count(n) > 1]
    results.append(ValidationResult(
        "Keine doppelten LKW-Blaetter",
        not duplicates,
        "OK" if not duplicates else f"Doppelt: {', '.join(duplicates)}",
    ))

    # 3: Reihenfolge entspricht dem Quellblatt
    order_ok = True
    for plan_vl, entries in groups.items():
        source_order = [
            normalize_lkw(_get(row, COL_PLAN_VL)) == plan_vl
            for row in rows
            if normalize_lkw(_get(row, COL_PLAN_VL)) == plan_vl
        ]
        if len(source_order) != len(entries):
            order_ok = False
    results.append(ValidationResult(
        "Reihenfolge entspricht Quellblatt",
        order_ok,
        "OK" if order_ok else "Reihenfolge weicht ab.",
    ))

    # 5: Gesamtgewicht = exakte Summe aus Spalte O
    weight_mismatches = []
    for plan_vl, entries in groups.items():
        expected_sum = sum((to_number(_get(row, COL_BRUTTO)) or 0)
                            for row in rows if normalize_lkw(_get(row, COL_PLAN_VL)) == plan_vl)
        actual_sum = sum((e.brutto or 0) for e in entries)
        if abs(expected_sum - actual_sum) > 1e-9:
            weight_mismatches.append(plan_vl)
    results.append(ValidationResult(
        "Gesamtgewicht stimmt mit Quelle ueberein",
        not weight_mismatches,
        "OK" if not weight_mismatches else f"Abweichung bei: {', '.join(weight_mismatches)}",
    ))

    # 6 + 7: 111444/VWPAL-Mengen stimmen, andere Ladungstraeger nicht enthalten
    pal_mismatches = []
    other_carriers_found = []
    for plan_vl, entries in groups.items():
        for entry in entries:
            expected_pal = process_pal(entry.pal_raw)
            if expected_pal != entry.pal:
                pal_mismatches.append(plan_vl)
            if entry.pal:
                for token in re.findall(r"\*([A-Za-z0-9]+)", entry.pal):
                    if token.upper() not in PAL_TOKENS:
                        other_carriers_found.append(plan_vl)
    results.append(ValidationResult(
        "PAL-Mengen (111444/VWPAL) korrekt extrahiert",
        not pal_mismatches,
        "OK" if not pal_mismatches else f"Abweichung bei: {', '.join(set(pal_mismatches))}",
    ))
    results.append(ValidationResult(
        "Keine anderen Ladungstraeger enthalten",
        not other_carriers_found,
        "OK" if not other_carriers_found else f"Fremde Ladungstraeger bei: {', '.join(set(other_carriers_found))}",
    ))

    # 8: alle Summenzeilen sind Formeln, keine Fehlerwerte
    formula_issues = []
    for plan_vl in groups.keys():
        sheet_name = _truck_sheet_name(plan_vl)
        ws = wb[sheet_name]
        n = 1 + len(groups[plan_vl])
        for row, col in ((n + 1, 7), (n + 2, 6), (n + 3, 6), (n + 4, 6)):
            cell = ws.cell(row=row, column=col)
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                formula_issues.append(f"{sheet_name}!{cell.coordinate}")
            elif cell.value.strip().upper() in ERROR_VALUES:
                formula_issues.append(f"{sheet_name}!{cell.coordinate}")
    results.append(ValidationResult(
        "Summen sind Formeln ohne Fehlerwerte",
        not formula_issues,
        "OK" if not formula_issues else f"Probleme bei: {', '.join(formula_issues)}",
    ))

    # 9: A:G sichtbar und korrekt formatiert; zusaetzliche Spalten (Helfer-
    # spalten fuer die PAL-Summenformeln) muessen ausgeblendet sein.
    layout_issues = []
    for plan_vl in groups.keys():
        sheet_name = _truck_sheet_name(plan_vl)
        ws = wb[sheet_name]
        problem = ws.max_column < 7
        for col_idx in range(8, ws.max_column + 1):
            if not ws.column_dimensions[get_column_letter(col_idx)].hidden:
                problem = True
                break
        if not problem:
            for col_letter, pt in COLUMN_WIDTHS_PT.items():
                expected = pt_to_excel_width(pt)
                actual = ws.column_dimensions[col_letter].width
                if actual is None or abs(actual - expected) > 0.05:
                    problem = True
                    break
        if problem:
            layout_issues.append(sheet_name)
    results.append(ValidationResult(
        "Nur Spalten A:G, Spaltenbreiten korrekt",
        not layout_issues,
        "OK" if not layout_issues else f"Probleme bei: {', '.join(layout_issues)}",
    ))

    # 10: Bestaetigungstexte nur ueber A:G verbunden
    merge_issues = []
    for plan_vl in groups.keys():
        sheet_name = _truck_sheet_name(plan_vl)
        ws = wb[sheet_name]
        merges = list(ws.merged_cells.ranges)
        if len(merges) != 3:
            merge_issues.append(sheet_name)
            continue
        for m in merges:
            if not (m.min_col == 1 and m.max_col == 7 and m.min_row == m.max_row):
                merge_issues.append(sheet_name)
                break
    results.append(ValidationResult(
        "Bestaetigungstexte ausschliesslich A:G verbunden",
        not merge_issues,
        "OK" if not merge_issues else f"Probleme bei: {', '.join(merge_issues)}",
    ))

    # 11: Druckbereich endet exakt in Unterschriftenzeile, passt auf eine Seite
    print_issues = []
    for plan_vl, entries in groups.items():
        sheet_name = _truck_sheet_name(plan_vl)
        ws = wb[sheet_name]
        n = 1 + len(entries)
        expected_signature_row = n + 12
        expected_area = f"A1:G{expected_signature_row}"
        actual_area = str(ws.print_area) if ws.print_area else ""
        actual_area = actual_area.split("!")[-1].replace("$", "")
        if actual_area != expected_area:
            print_issues.append(sheet_name)
            continue
        if ws.page_setup.fitToWidth != 1:
            print_issues.append(sheet_name)
    results.append(ValidationResult(
        "Druckbereich/Skalierung korrekt",
        not print_issues,
        "OK" if not print_issues else f"Probleme bei: {', '.join(print_issues)}",
    ))

    return results


# ---------------------------------------------------------------------------
# Schritt 4: Planungsabgleich (optional, nur wenn eine Planungsdatei
# hochgeladen wurde)
# ---------------------------------------------------------------------------

@dataclass
class AbweichungEntry:
    abladestelle: str
    erwartet_vwpal: int
    erwartet_111444: int
    tatsaechlich_vwpal: int
    tatsaechlich_111444: int
    beschreibung: str


@dataclass
class PlanungAbgleichResult:
    positions_total: int = 0
    positions_gefuellt: int = 0
    positions_ignoriert: int = 0
    positions_ohne_treffer: int = 0
    ambiguous_abladestellen: Dict[str, List[str]] = field(default_factory=dict)
    abweichungen: List[AbweichungEntry] = field(default_factory=list)


def _normalize_palet_type(value: Any) -> Optional[str]:
    """Normalisiert eine PALET-Angabe aus der Planung auf "VWPAL"/"111444"
    oder None, falls es sich um eine andere Art handelt (z.B. "-", "99C159")
    oder keine Angabe vorhanden ist.
    """
    if value is None:
        return None
    text = str(value).strip().upper()
    if text == "VWPAL":
        return "VWPAL"
    if text == "111444":
        return "111444"
    return None


def _clone_sheet_exact(source_ws: Worksheet, target_wb: Workbook, target_name: str) -> Worksheet:
    """Kopiert ein komplettes Blatt 1:1 (Werte, Formatierung, Spaltenbreiten,
    Zeilenhoehen, verbundene Zellen) in eine neue Arbeitsmappe.
    """
    target_ws = target_wb.create_sheet(target_name)

    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width
        target_ws.column_dimensions[col_letter].hidden = dim.hidden

    for row_idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = dim.height

    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    return target_ws


def _build_abladestelle_plan_vl_map(
    groups: "OrderedDict[str, List[TruckEntry]]",
) -> Tuple[Dict[str, str], Dict[str, List[str]]]:
    """Ermittelt je Abladestelle den (haeufigsten) LKW aus den Avis-Daten.

    Kommt eine Abladestelle bei mehreren unterschiedlichen LKW vor, wird der
    haeufigste genommen und die Abladestelle zusaetzlich als mehrdeutig
    vermerkt (informativ, kein Fehler).
    """
    counters: Dict[str, Counter] = {}
    for plan_vl, entries in groups.items():
        for entry in entries:
            kapi = normalize_lkw(entry.abladestelle)
            if kapi is None:
                continue
            counters.setdefault(kapi, Counter())[plan_vl] += 1

    mapping: Dict[str, str] = {}
    ambiguous: Dict[str, List[str]] = {}
    for kapi, counter in counters.items():
        mapping[kapi] = counter.most_common(1)[0][0]
        if len(counter) > 1:
            ambiguous[kapi] = sorted(counter.keys())
    return mapping, ambiguous


def _collect_avis_actual(groups: "OrderedDict[str, List[TruckEntry]]") -> Dict[str, Dict[str, int]]:
    """Summiert je Abladestelle die tatsaechlichen VWPAL/111444-Mengen aus
    den (bereits gefilterten) Avis-Daten - unabhaengig vom LKW.
    """
    actual: Dict[str, Dict[str, int]] = {}
    for entries in groups.values():
        for entry in entries:
            kapi = normalize_lkw(entry.abladestelle)
            if kapi is None:
                continue
            bucket = actual.setdefault(kapi, {"VWPAL": 0, "111444": 0})
            bucket["VWPAL"] += _extract_qty(entry.pal, "VWPAL")
            bucket["111444"] += _extract_qty(entry.pal, "111444")
    return actual


def _collect_planung_expected(source_ws: Worksheet) -> Dict[str, Dict[str, int]]:
    """Summiert je Abladestelle die in der Planung erwarteten VWPAL/111444-
    Mengen (andere PALET-Arten wie "-" oder "99C159" werden ignoriert).
    """
    expected: Dict[str, Dict[str, int]] = {}
    for row_idx in range(PLANUNG_DATA_START_ROW, source_ws.max_row + 1):
        kapi = normalize_lkw(source_ws.cell(row=row_idx, column=PLANUNG_COL_KAPI).value)
        if kapi is None:
            continue
        palet_type = _normalize_palet_type(source_ws.cell(row=row_idx, column=PLANUNG_COL_PALET).value)
        if palet_type is None:
            continue
        qty = to_number(source_ws.cell(row=row_idx, column=PLANUNG_COL_PALET_SAYISI).value) or 0
        bucket = expected.setdefault(kapi, {"VWPAL": 0, "111444": 0})
        bucket[palet_type] += qty
    return expected


def _describe_abweichung(exp: Dict[str, int], act: Dict[str, int]) -> str:
    exp_v, exp_1 = exp["VWPAL"], exp["111444"]
    act_v, act_1 = act["VWPAL"], act["111444"]

    # Klassischer "falscher Typ"-Fall: eine Art fehlt komplett, dafuer ist
    # (mindestens) die erwartete Menge der jeweils anderen Art vorhanden.
    if exp_v > 0 and exp_1 == 0 and act_v == 0 and act_1 >= exp_v:
        return f"Falscher Typ: {exp_v}*VWPAL erwartet, aber 111444 geliefert"
    if exp_1 > 0 and exp_v == 0 and act_1 == 0 and act_v >= exp_1:
        return f"Falscher Typ: {exp_1}*111444 erwartet, aber VWPAL geliefert"

    parts = []
    if exp_v != act_v:
        if act_v == 0:
            parts.append(f"VWPAL fehlt komplett (erwartet {exp_v})")
        else:
            parts.append(f"VWPAL: erwartet {exp_v}, tatsaechlich {act_v}")
    if exp_1 != act_1:
        if act_1 == 0:
            parts.append(f"111444 fehlt komplett (erwartet {exp_1})")
        else:
            parts.append(f"111444: erwartet {exp_1}, tatsaechlich {act_1}")
    return "; ".join(parts) if parts else "Abweichung"


def _write_abweichungen_sheet(wb: Workbook, abweichungen: List[AbweichungEntry]) -> Worksheet:
    ws = wb.create_sheet(ABWEICHUNGEN_SHEET_NAME)
    headers = [
        "Abladestelle",
        "Erwartet VWPAL",
        "Erwartet 111444",
        "Tatsaechlich VWPAL",
        "Tatsaechlich 111444",
        "Abweichung",
    ]
    ws.append(headers)
    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_COLOR)
    header_fill = PatternFill(fill_type="solid", start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for entry in abweichungen:
        ws.append([
            entry.abladestelle,
            entry.erwartet_vwpal,
            entry.erwartet_111444,
            entry.tatsaechlich_vwpal,
            entry.tatsaechlich_111444,
            entry.beschreibung,
        ])

    if not abweichungen:
        ws.cell(row=2, column=1, value="Keine Abweichungen gefunden.")

    widths = {"A": 20, "B": 16, "C": 16, "D": 18, "E": 18, "F": 55}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    return ws


def add_planung_reconciliation(
    wb: Workbook,
    planung_file: Any,
    groups: "OrderedDict[str, List[TruckEntry]]",
) -> PlanungAbgleichResult:
    """Liest das Blatt "Planung fuer Staplerfahrer" aus der hochgeladenen
    Planungsdatei, kopiert es 1:1 in die Ausgabe-Arbeitsmappe, befuellt dort
    Spalte J (LKW) anhand der Abladestelle-zu-LKW-Zuordnung aus den
    Avis-Daten und erstellt ein Blatt "Abweichungen" mit allen Unterschieden
    zwischen Planung und tatsaechlichen Avis-Mengen (in beide Richtungen).
    """
    source_wb = load_workbook(planung_file, data_only=True)
    if PLANUNG_SHEET_NAME not in source_wb.sheetnames:
        raise ValueError(
            f"Das Blatt '{PLANUNG_SHEET_NAME}' wurde in der Planungsdatei nicht gefunden."
        )
    source_ws = source_wb[PLANUNG_SHEET_NAME]

    target_ws = _clone_sheet_exact(source_ws, wb, PLANUNG_SHEET_NAME)

    plan_vl_map, ambiguous = _build_abladestelle_plan_vl_map(groups)
    avis_actual = _collect_avis_actual(groups)

    result = PlanungAbgleichResult(ambiguous_abladestellen=ambiguous)
    planung_abladestellen: set = set()

    for row_idx in range(PLANUNG_DATA_START_ROW, source_ws.max_row + 1):
        kapi = normalize_lkw(source_ws.cell(row=row_idx, column=PLANUNG_COL_KAPI).value)
        if kapi is None:
            continue
        result.positions_total += 1
        if kapi.upper() in IGNORED_ABLADESTELLEN:
            result.positions_ignoriert += 1
            continue
        planung_abladestellen.add(kapi)
        plan_vl = plan_vl_map.get(kapi)
        if plan_vl is not None:
            target_ws.cell(row=row_idx, column=PLANUNG_COL_LKW, value=plan_vl)
            result.positions_gefuellt += 1
        else:
            target_ws.cell(row=row_idx, column=PLANUNG_COL_LKW).value = None
            result.positions_ohne_treffer += 1

    expected = _collect_planung_expected(source_ws)

    # Richtung 1: Planung -> Avis (fehlende/falsche/abweichende Mengen).
    for kapi, exp in expected.items():
        if kapi.upper() in IGNORED_ABLADESTELLEN:
            continue
        act = avis_actual.get(kapi, {"VWPAL": 0, "111444": 0})
        if exp["VWPAL"] == act["VWPAL"] and exp["111444"] == act["111444"]:
            continue
        result.abweichungen.append(AbweichungEntry(
            abladestelle=kapi,
            erwartet_vwpal=exp["VWPAL"],
            erwartet_111444=exp["111444"],
            tatsaechlich_vwpal=act["VWPAL"],
            tatsaechlich_111444=act["111444"],
            beschreibung=_describe_abweichung(exp, act),
        ))

    # Richtung 2: Avis -> Planung (Abladestellen, die es in den Avis-Daten
    # gibt, aber gar nicht in der Planung vorkommen).
    for kapi, act in avis_actual.items():
        if kapi.upper() in IGNORED_ABLADESTELLEN or kapi in planung_abladestellen:
            continue
        result.abweichungen.append(AbweichungEntry(
            abladestelle=kapi,
            erwartet_vwpal=0,
            erwartet_111444=0,
            tatsaechlich_vwpal=act["VWPAL"],
            tatsaechlich_111444=act["111444"],
            beschreibung="Nicht in der Planung enthalten (nur in den Avis-Daten gefunden).",
        ))

    _write_abweichungen_sheet(wb, result.abweichungen)

    source_wb.close()
    return result


# ---------------------------------------------------------------------------
# Gesamtablauf
# ---------------------------------------------------------------------------

@dataclass
class ProcessingResult:
    workbook: Workbook
    summaries: List[TruckSheetSummary]
    validation: List[ValidationResult]
    planung: Optional[PlanungAbgleichResult] = None

    @property
    def is_valid(self) -> bool:
        return all(v.passed for v in self.validation)


def generate_workbook(files: Sequence[Any], planung_file: Any = None) -> ProcessingResult:
    header, rows = merge_source_files(files)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    write_source_sheet(wb, header, rows)

    groups = build_truck_groups(rows)

    summaries: List[TruckSheetSummary] = []
    for plan_vl, entries in groups.items():
        _, summary = build_truck_sheet(wb, plan_vl, entries)
        summaries.append(summary)

    validation = validate(header, rows, groups, wb)

    planung_result = None
    if planung_file is not None:
        planung_result = add_planung_reconciliation(wb, planung_file, groups)

    return ProcessingResult(workbook=wb, summaries=summaries, validation=validation, planung=planung_result)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Ladelisten-Tool (CLI)")
    parser.add_argument("inputs", nargs="+", help="Eine oder mehrere Excel-Quelldateien")
    parser.add_argument("-o", "--output", default="Ladelisten.xlsx", help="Pfad der Ausgabedatei")
    parser.add_argument("--planung", default=None, help="Optionale Planungsdatei (Planung fuer Staplerfahrer)")
    args = parser.parse_args(argv)

    result = generate_workbook(args.inputs, planung_file=args.planung)

    for s in result.summaries:
        print(f"{s.sheet_name}: {s.position_count} Positionen, "
              f"Gesamtgewicht={s.gesamtgewicht}, 111444={s.total_111444}, VWPAL={s.total_vwpal}")

    if result.planung is not None:
        p = result.planung
        print(f"\nPlanungsabgleich: {p.positions_gefuellt} befuellt, "
              f"{p.positions_ignoriert} ignoriert, {p.positions_ohne_treffer} ohne Treffer, "
              f"{len(p.abweichungen)} Abweichung(en).")

    failed = [v for v in result.validation if not v.passed]
    for v in result.validation:
        status = "OK" if v.passed else "FEHLER"
        print(f"[{status}] {v.check}: {v.message}")

    if failed:
        print(f"\n{len(failed)} Validierung(en) fehlgeschlagen. Datei wird nicht gespeichert.", file=sys.stderr)
        return 1

    result.workbook.save(args.output)
    print(f"\nGespeichert: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
