import io
import sys
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import ladeliste_logic as ll


# ---------------------------------------------------------------------------
# PAL-Regex-Logik
# ---------------------------------------------------------------------------

def test_pal_example_1_drei_ladungstraeger():
    text = "10*VWPAL / 19*111444 / 600*003147"
    assert ll.process_pal(text) == "10*VWPAL / 19*111444"


def test_pal_example_2_nur_111444():
    text = "3*111444 / 3*001006 / 70*003147"
    assert ll.process_pal(text) == "3*111444"


def test_pal_example_3_kein_treffer():
    text = "Dein"
    assert ll.process_pal(text) == ""


def test_pal_none_wert():
    assert ll.process_pal(None) == ""


def test_pal_leerer_text():
    assert ll.process_pal("") == ""


def test_pal_reihenfolge_111444_zuerst():
    text = "7*111444 / 4*VWPAL"
    assert ll.process_pal(text) == "7*111444 / 4*VWPAL"


def test_pal_nur_vwpal():
    text = "5*VWPAL"
    assert ll.process_pal(text) == "5*VWPAL"


def test_pal_case_insensitive():
    text = "8*vwpal / 2*111444"
    assert ll.process_pal(text) == "8*VWPAL / 2*111444"


def test_pal_beliebige_leerzeichen_um_stern_und_slash():
    text = "10 *  VWPAL   /   19*  111444"
    assert ll.process_pal(text) == "10*VWPAL / 19*111444"


def test_pal_ohne_leerzeichen_um_stern():
    text = "10*VWPAL/19*111444"
    assert ll.process_pal(text) == "10*VWPAL / 19*111444"


def test_pal_mehrfaches_vorkommen_wird_summiert():
    text = "5*VWPAL / 2*111444 / 3*VWPAL"
    assert ll.process_pal(text) == "8*VWPAL / 2*111444"


def test_pal_ignoriert_aehnliche_aber_andere_ladungstraeger():
    text = "100*111445 / 3*VWPALX / 4*111444"
    assert ll.process_pal(text) == "4*111444"


def test_pal_haelt_reale_mehrfach_ladungstraeger_wie_in_produktivdaten():
    # Beispiel aus einer echten Ladeliste: mehrere fremde Ladungstraeger
    # zwischen den beiden relevanten Arten, mit unregelmaessigen Abstaenden.
    text = "   6*111444  /   10*VWPAL   /   6*001006  /  10*001210  / 461*003147  "
    assert ll.process_pal(text) == "6*111444 / 10*VWPAL"


# ---------------------------------------------------------------------------
# Zusammenfuehren mehrerer Rohdateien
# ---------------------------------------------------------------------------
# Die Rohdateien haben die volle Spaltenstruktur (mind. bis Spalte O). Das
# Zusammenfuehren bildet jede Zeile sofort auf die 7 Kernspalten ab:
# Abholtag/Empfaenger Name/Empfaenger Ort/Abladestelle/Plan VL/PAL/Brutto.
# Massgeblich fuer die LKW-Zuordnung ist Spalte K "Plan VL" - NICHT Spalte L
# "LKW VL", die in echten Exporten haeufig leer ist.

RAW_HEADER = [f"col{i}" for i in range(19)]


def _raw_row(plan_vl="GEVC62-1", pal="1*VWPAL", brutto=100, name="Empfaenger", lkw_vl=None):
    row = [None] * 19
    row[ll.RAW_COL_ABHOLTAG] = "01.01.2026"
    row[ll.RAW_COL_EMPF_NAME] = name
    row[ll.RAW_COL_EMPF_ORT] = "Ort"
    row[ll.RAW_COL_ABLADESTELLE] = "Stelle"
    row[ll.RAW_COL_PLAN_VL] = plan_vl
    row[11] = lkw_vl  # Spalte L "LKW VL" - bewusst haeufig leer, darf nicht verwendet werden
    row[ll.RAW_COL_PAL_QUELLE] = pal
    row[ll.RAW_COL_BRUTTO] = brutto
    return row


def _make_source_workbook_bytes(rows, header=None):
    wb = Workbook()
    ws = wb.active
    ws.title = ll.SOURCE_SHEET_NAME
    ws.append(header or RAW_HEADER)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_merge_haelt_reihenfolge_ueber_mehrere_dateien():
    rows1 = [_raw_row(name="A1"), _raw_row(name="A2"), _raw_row(name="A3")]
    rows2 = [_raw_row(name="B1"), _raw_row(name="B2")]
    file1 = _make_source_workbook_bytes(rows1)
    file2 = _make_source_workbook_bytes(rows2)

    header, merged = ll.merge_source_files([file1, file2])

    names = [row[ll.COL_EMPF_NAME] for row in merged]
    assert names == ["A1", "A2", "A3", "B1", "B2"]


def test_merge_verwendet_immer_die_festen_7_kernspalten_als_kopfzeile():
    file1 = _make_source_workbook_bytes([_raw_row(name="A1")], header=RAW_HEADER)
    file2 = _make_source_workbook_bytes(
        [_raw_row(name="B1")], header=[f"anders{i}" for i in range(19)]
    )

    header, merged = ll.merge_source_files([file1, file2])

    assert header == ll.TRUCK_SHEET_HEADERS
    assert len(merged) == 2


def test_merge_bildet_zeilen_auf_7_spalten_ab():
    file1 = _make_source_workbook_bytes([_raw_row(name="A1", plan_vl="GEVC62-1", brutto=123.4)])
    _, merged = ll.merge_source_files([file1])

    assert len(merged[0]) == 7
    assert merged[0][ll.COL_EMPF_NAME] == "A1"
    assert merged[0][ll.COL_PLAN_VL] == "GEVC62-1"
    assert merged[0][ll.COL_BRUTTO] == 123.4


def test_merge_ohne_sortierung_oder_dedupe():
    rows1 = [_raw_row(name="Gleich"), _raw_row(name="Gleich")]
    file1 = _make_source_workbook_bytes(rows1)

    header, merged = ll.merge_source_files([file1])

    assert len(merged) == 2


def test_merge_ohne_dateien_wirft_fehler():
    try:
        ll.merge_source_files([])
        assert False, "Erwartete ValueError"
    except ValueError:
        pass


# ---------------------------------------------------------------------------
# LKW-Gruppierung (nach Spalte "Plan VL" der bereits zusammengefuehrten Zeilen)
# ---------------------------------------------------------------------------

def _merged_row(plan_vl="GEVC62-1", pal="1*VWPAL", brutto=100, name="Empfaenger"):
    row = [None] * 7
    row[ll.COL_ABHOLTAG] = "01.01.2026"
    row[ll.COL_EMPF_NAME] = name
    row[ll.COL_EMPF_ORT] = "Ort"
    row[ll.COL_ABLADESTELLE] = "Stelle"
    row[ll.COL_PLAN_VL] = plan_vl
    row[ll.COL_PAL_QUELLE] = pal
    row[ll.COL_BRUTTO] = brutto
    return row


def test_build_truck_groups_ignoriert_leeren_plan_vl():
    rows = [
        _merged_row(plan_vl="LKW1"),
        _merged_row(plan_vl=None),
        _merged_row(plan_vl="   "),
        _merged_row(plan_vl="LKW2"),
    ]
    groups = ll.build_truck_groups(rows)
    assert list(groups.keys()) == ["LKW1", "LKW2"]
    assert len(groups["LKW1"]) == 1
    assert len(groups["LKW2"]) == 1


def test_build_truck_groups_erhaelt_reihenfolge_je_gruppe():
    rows = [
        _merged_row(plan_vl="LKW1", name="Erste"),
        _merged_row(plan_vl="LKW2", name="Zweite"),
        _merged_row(plan_vl="LKW1", name="Dritte"),
    ]
    groups = ll.build_truck_groups(rows)
    names = [e.empf_name for e in groups["LKW1"]]
    assert names == ["Erste", "Dritte"]


def test_build_truck_groups_behandelt_suffix_als_eigenstaendige_kennung():
    # "BOHDT978-1" und "BOHDT978-2" sind unterschiedliche Ladungen desselben
    # LKW und muessen in getrennte Gruppen/Blaetter gehen (kein Abschneiden
    # des "-N"-Zusatzes).
    rows = [
        _merged_row(plan_vl="BOHDT978-1", name="A"),
        _merged_row(plan_vl="BOHDT978-2", name="B"),
    ]
    groups = ll.build_truck_groups(rows)
    assert set(groups.keys()) == {"BOHDT978-1", "BOHDT978-2"}


# ---------------------------------------------------------------------------
# End-to-End: generierte Arbeitsmappe besteht alle Validierungen
# ---------------------------------------------------------------------------

def test_generate_workbook_end_to_end_validiert_fehlerfrei():
    rows = [
        _raw_row(plan_vl="LKW1", pal="10*VWPAL / 19*111444 / 600*003147", brutto=9350.3, name="A"),
        _raw_row(plan_vl="LKW1", pal="3*111444 / 3*001006 / 70*003147", brutto=100.5, name="B"),
        _raw_row(plan_vl="LKW2", pal="Dein", brutto=50, name="C"),
    ]
    file1 = _make_source_workbook_bytes(rows)

    result = ll.generate_workbook([file1])

    assert result.is_valid, [v.message for v in result.validation if not v.passed]
    assert {s.sheet_name for s in result.summaries} == {"LKW LKW1", "LKW LKW2"}

    lkw1 = next(s for s in result.summaries if s.sheet_name == "LKW LKW1")
    assert lkw1.position_count == 2
    assert lkw1.total_111444 == 22
    assert lkw1.total_vwpal == 10
    assert abs(lkw1.gesamtgewicht - 9450.8) < 1e-9


def test_generate_workbook_gruppiert_nach_plan_vl_auch_wenn_lkw_vl_spalte_leer_ist():
    # Regressionstest fuer den realen Fehler: Spalte L "LKW VL" ist in
    # Produktivdaten haeufig leer, Spalte K "Plan VL" ist aber immer gefuellt
    # und muss fuer die Zuordnung verwendet werden.
    rows = [
        _raw_row(plan_vl="CZ2A", lkw_vl=None, brutto=100, name="A"),
        _raw_row(plan_vl="CZ2A", lkw_vl=None, brutto=200, name="B"),
    ]
    file1 = _make_source_workbook_bytes(rows)

    result = ll.generate_workbook([file1])

    assert result.is_valid, [v.message for v in result.validation if not v.passed]
    assert {s.sheet_name for s in result.summaries} == {"LKW CZ2A"}
    summary = result.summaries[0]
    assert summary.position_count == 2
    assert abs(summary.gesamtgewicht - 300) < 1e-9


# ---------------------------------------------------------------------------
# Planungsabgleich ("Planung fuer Staplerfahrer" + "Abweichungen")
# ---------------------------------------------------------------------------

def _make_planung_workbook_bytes(data_rows):
    """Baut eine minimale Planungsdatei mit dem Blatt "Planung fuer
    Staplerfahrer" nach), Kopfzeile in Zeile 4, Daten ab Zeile 5, Spalten
    B=KAPI, H=PALET SAYISI, I=PALET, J=LKW (wie im echten Format).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = ll.PLANUNG_SHEET_NAME
    ws.append([None] * 10)  # Zeile 1 leer
    ws.append([None] * 10)  # Zeile 2
    ws.append([None] * 10)  # Zeile 3
    ws.append(["Werk", "KAPI", "MALZEME", "ADET", "KUTU ICI", "KLT SAYISI",
                "NAKLIYE NR", "PALET SAYISI", "PALET", "LKW"])  # Zeile 4
    for kapi, palet_sayisi, palet, lkw in data_rows:
        row = [None] * 10
        row[ll.PLANUNG_COL_KAPI - 1] = kapi
        row[ll.PLANUNG_COL_PALET_SAYISI - 1] = palet_sayisi
        row[ll.PLANUNG_COL_PALET - 1] = palet
        row[ll.PLANUNG_COL_LKW - 1] = lkw
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_planung_befuellt_lkw_spalte_ueber_abladestelle_match():
    avis_rows = [
        _raw_row(plan_vl="CZ2A", pal="2*VWPAL", brutto=10, name="A"),
    ]
    # Abladestelle in _raw_row ist immer "Stelle" -> als KAPI in der Planung nutzen.
    avis_file = _make_source_workbook_bytes(avis_rows)
    planung_file = _make_planung_workbook_bytes([
        ("Stelle", 2, "VWPAL", None),  # LKW zunaechst leer, wie im echten Fall
    ])

    result = ll.generate_workbook([avis_file], planung_file=planung_file)

    ws = result.workbook[ll.PLANUNG_SHEET_NAME]
    assert ws.cell(row=5, column=ll.PLANUNG_COL_LKW).value == "CZ2A"
    assert result.planung.positions_gefuellt == 1
    assert result.planung.positions_ohne_treffer == 0
    assert result.planung.positions_ignoriert == 0


def test_planung_ignoriert_bekannte_abladestellen_ohne_avis_bezug():
    avis_rows = [_raw_row(plan_vl="CZ2A", pal="1*VWPAL", brutto=10, name="A")]
    avis_file = _make_source_workbook_bytes(avis_rows)
    planung_file = _make_planung_workbook_bytes([
        ("BAU90", 1, "99C159", "AlteHandeintragung"),
    ])

    result = ll.generate_workbook([avis_file], planung_file=planung_file)

    ws = result.workbook[ll.PLANUNG_SHEET_NAME]
    # Ignorierte Abladestelle bleibt unangetastet (nicht ueberschrieben).
    assert ws.cell(row=5, column=ll.PLANUNG_COL_LKW).value == "AlteHandeintragung"
    assert result.planung.positions_ignoriert == 1
    assert result.planung.positions_gefuellt == 0
    # Ignorierte Abladestellen duerfen nicht in den Abweichungen auftauchen.
    assert all(a.abladestelle != "BAU90" for a in result.planung.abweichungen)


def test_planung_ohne_avis_treffer_bleibt_leer_und_wird_gemeldet():
    avis_rows = [_raw_row(plan_vl="CZ2A", pal="1*VWPAL", brutto=10, name="A")]
    avis_file = _make_source_workbook_bytes(avis_rows)
    planung_file = _make_planung_workbook_bytes([
        ("UnbekannteStelle", 3, "111444", "AlteHandeintragung"),
    ])

    result = ll.generate_workbook([avis_file], planung_file=planung_file)

    ws = result.workbook[ll.PLANUNG_SHEET_NAME]
    assert ws.cell(row=5, column=ll.PLANUNG_COL_LKW).value is None
    assert result.planung.positions_ohne_treffer == 1
    abweichung = next(a for a in result.planung.abweichungen if a.abladestelle == "UnbekannteStelle")
    assert "fehlt komplett" in abweichung.beschreibung


def test_planung_erkennt_falschen_typ():
    # Planung erwartet VWPAL, Avis liefert stattdessen 111444 an derselben Abladestelle.
    avis_rows = [_raw_row(plan_vl="CZ2A", pal="2*111444", brutto=10, name="A")]
    avis_file = _make_source_workbook_bytes(avis_rows)
    planung_file = _make_planung_workbook_bytes([
        ("Stelle", 2, "VWPAL", None),
    ])

    result = ll.generate_workbook([avis_file], planung_file=planung_file)

    abweichung = next(a for a in result.planung.abweichungen if a.abladestelle == "Stelle")
    assert "Falscher Typ" in abweichung.beschreibung


def test_planung_erkennt_unerwartete_abladestelle_nur_in_avis():
    avis_rows = [_raw_row(plan_vl="CZ2A", pal="1*VWPAL", brutto=10, name="A")]
    avis_file = _make_source_workbook_bytes(avis_rows)
    # Planungsdatei enthaelt die Abladestelle "Stelle" ueberhaupt nicht.
    planung_file = _make_planung_workbook_bytes([
        ("AndereStelle", 1, "VWPAL", "CZ2A"),
    ])

    result = ll.generate_workbook([avis_file], planung_file=planung_file)

    abweichung = next(a for a in result.planung.abweichungen if a.abladestelle == "Stelle")
    assert "Nicht in der Planung enthalten" in abweichung.beschreibung


def test_planung_ohne_abweichung_wenn_mengen_uebereinstimmen():
    avis_rows = [_raw_row(plan_vl="CZ2A", pal="2*VWPAL", brutto=10, name="A")]
    avis_file = _make_source_workbook_bytes(avis_rows)
    planung_file = _make_planung_workbook_bytes([
        ("Stelle", 2, "VWPAL", None),
    ])

    result = ll.generate_workbook([avis_file], planung_file=planung_file)

    assert result.planung.abweichungen == []


def test_planung_blatt_erhaelt_format_und_zusatzblatt_wird_erzeugt():
    avis_rows = [_raw_row(plan_vl="CZ2A", pal="1*VWPAL", brutto=10, name="A")]
    avis_file = _make_source_workbook_bytes(avis_rows)
    planung_file = _make_planung_workbook_bytes([("Stelle", 1, "VWPAL", None)])

    result = ll.generate_workbook([avis_file], planung_file=planung_file)

    assert ll.PLANUNG_SHEET_NAME in result.workbook.sheetnames
    assert ll.ABWEICHUNGEN_SHEET_NAME in result.workbook.sheetnames
    ws = result.workbook[ll.PLANUNG_SHEET_NAME]
    # Kopfzeile (Zeile 4) bleibt unveraendert erhalten.
    assert ws.cell(row=4, column=ll.PLANUNG_COL_KAPI).value == "KAPI"


def test_generate_workbook_ohne_planungsdatei_hat_kein_planung_ergebnis():
    avis_rows = [_raw_row(plan_vl="CZ2A", pal="1*VWPAL", brutto=10, name="A")]
    avis_file = _make_source_workbook_bytes(avis_rows)

    result = ll.generate_workbook([avis_file])

    assert result.planung is None
    assert ll.PLANUNG_SHEET_NAME not in result.workbook.sheetnames
    assert ll.ABWEICHUNGEN_SHEET_NAME not in result.workbook.sheetnames
