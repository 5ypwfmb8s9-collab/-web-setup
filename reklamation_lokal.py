"""Lokale Erfassung von Ausfallfracht-/Storno-Reklamationen per Drag & Drop.

Alternative zum Power-Automate-Weg (siehe reklamation_logic.py /
REKLAMATIONEN_SETUP.md, weiterhin nutzbar falls gewuenscht): PDFs werden
manuell per Drag & Drop hochgeladen, landen als echte Datei im
PDF-Unterordner und werden als Zeile in einer Excel-Datei erfasst - beide
in einem frei waehlbaren Basisordner (z.B. einem SharePoint/OneDrive-
synchronisierten Ordner), kein Power-Automate/SharePoint-Setup noetig.

Excel-Spalten und Blattname basieren auf der vom Nutzer vorgegebenen
Vorlage (VW_Reklamationen.xlsx, Blatt "Tabelle1"), erweitert um
Abholtag, Firma, Betrag, Zugewiesen und Referenznummern: Eingangsdatum,
Abholtag, Firma, Absender, Betreff, Dateiname_PDF, OneDrive_Pfad,
Betrag, Zugewiesen, Status, Pruefdatum, Ergebnis, Bemerkung,
Referenznummern.
"""

from __future__ import annotations

import datetime
import io
import os
import re
import shutil
from typing import Any, Dict, List, Optional, Tuple

import pdfplumber
from openpyxl import Workbook, load_workbook

SHEET_NAME = "Tabelle1"
EXCEL_COLUMNS = [
    "Eingangsdatum",
    "Abholtag",
    "Firma",
    "Absender",
    "Betreff",
    "Dateiname_PDF",
    "OneDrive_Pfad",
    "Betrag",
    "Zugewiesen",
    "Status",
    "Prüfdatum",
    "Ergebnis",
    "Bemerkung",
    "Referenznummern",
]
EXCEL_SPALTENBREITEN = [16, 14, 16, 22, 26, 18, 45, 14, 16, 16, 14, 14, 30, 24]

ZUGEWIESEN_OPTIONEN = ["", "Murat Kurt", "Okan Kocak", "Alperen Konar", "Levin Akarcay"]
ERGEBNIS_OPTIONEN = ["", "Berechtigt", "Unberechtigt"]

PFAD_CONFIG_DATEI = ".reklamationen_basisordner.txt"
STANDARD_BASISORDNER = (
    r"C:\Users\okan.kocak\norm-fasteners.com.tr"
    r"\Norm Fasteners Germany - VWAI Projekt\VW_Reklamationen"
)
EXCEL_DATEINAME = "VW_Reklamationen.xlsx"
PDF_UNTERORDNER = "Rechnungen_PDF"

ARCHIV_CONFIG_DATEI = ".reklamationen_archivordner.txt"
# {jahr} wird automatisch durch das Jahr des Abholtags ersetzt (siehe
# archiv_monatsordner) - so muss der Pfad beim Jahreswechsel nicht von
# Hand angepasst werden, solange sich nur die Jahreszahl im Ordnernamen
# aendert.
STANDARD_ARCHIV_BASISORDNER = (
    r"C:\Users\okan.kocak\norm-fasteners.com.tr"
    r"\Norm Fasteners Germany - {jahr}"
)
FALLORDNER_PRAEFIX = "Fall-"

PLANUNGSDATEI_TYPEN = {
    "Planung": "VW_Planung",
    "Ladeliste": "VW_Ladeliste",
    "Avisierung": "VW_Avisierung",
}

ABSENDER_DUVENBECK = "Duvenbeck"
BETREFF_AUSFALLFRACHT = "Ausfallfracht Rechnung"
BETREFF_STORNO = "Ausfallfracht Storno"

_DATUM_MUSTER = re.compile(r"(20\d{2})(\d{2})(\d{2})")


def lade_gespeicherten_basisordner() -> str:
    """Liest den zuletzt gespeicherten Basisordner, falls vorhanden."""
    try:
        with open(PFAD_CONFIG_DATEI, "r", encoding="utf-8") as f:
            pfad = f.read().strip()
            if pfad:
                return pfad
    except FileNotFoundError:
        pass
    return STANDARD_BASISORDNER


def speichere_basisordner(pfad: str) -> None:
    """Merkt sich den Basisordner dauerhaft fuer kuenftige Programmstarts."""
    with open(PFAD_CONFIG_DATEI, "w", encoding="utf-8") as f:
        f.write(pfad.strip())


def lade_gespeicherten_archivordner() -> str:
    """Liest den zuletzt gespeicherten Archiv-Basisordner, falls vorhanden."""
    try:
        with open(ARCHIV_CONFIG_DATEI, "r", encoding="utf-8") as f:
            pfad = f.read().strip()
            if pfad:
                return pfad
    except FileNotFoundError:
        pass
    return STANDARD_ARCHIV_BASISORDNER


def speichere_archivordner(pfad: str) -> None:
    """Merkt sich den Archiv-Basisordner dauerhaft fuer kuenftige Programmstarts."""
    with open(ARCHIV_CONFIG_DATEI, "w", encoding="utf-8") as f:
        f.write(pfad.strip())


def excel_pfad(basisordner: str) -> str:
    return os.path.join(basisordner, EXCEL_DATEINAME)


def pdf_ordner(basisordner: str) -> str:
    return os.path.join(basisordner, PDF_UNTERORDNER)


def guess_eingangsdatum(dateiname: str) -> str:
    """Rät ein Datum (TT.MM.JJJJ) aus einem JJJJMMTT-Muster im Dateinamen,
    faellt sonst auf das heutige Datum zurueck."""
    match = _DATUM_MUSTER.search(dateiname)
    if match:
        jahr, monat, tag = match.groups()
        if 1 <= int(monat) <= 12 and 1 <= int(tag) <= 31:
            return f"{tag}.{monat}.{jahr}"
    return datetime.date.today().strftime("%d.%m.%Y")


def guess_absender_betreff(dateiname: str) -> Dict[str, str]:
    """Raet Absender/Betreff anhand des Dateinamens (Storno vs. Rechnung)."""
    betreff = BETREFF_STORNO if "storno" in dateiname.lower() else BETREFF_AUSFALLFRACHT
    return {"Absender": ABSENDER_DUVENBECK, "Betreff": betreff}


def _wert_unter_label(
    words: List[Dict[str, Any]],
    label: Dict[str, Any],
    x_toleranz: float = 6.0,
    y_min: float = 2.0,
    y_max: float = 25.0,
) -> Optional[str]:
    """Sucht das Wort direkt unter einem Label (gleiche linke Kante, naechste Zeile)."""
    kandidaten = [
        w
        for w in words
        if label["bottom"] + y_min <= w["top"] <= label["bottom"] + y_max
        and abs(w["x0"] - label["x0"]) <= x_toleranz
    ]
    if not kandidaten:
        return None
    kandidaten.sort(key=lambda w: w["top"])
    return kandidaten[0]["text"]


def extrahiere_beleg_und_datum(
    pdf_bytes: bytes,
) -> Tuple[Optional[str], Optional[str]]:
    """Liest Beleg-Nr. und Datum aus dem Kasten "Bei Zahlung bitte angeben"
    (erste Seite): der Wert steht jeweils direkt unter dem gleich weit
    links stehenden Label. Gibt (None, None) zurueck, wenn das Layout
    nicht erkannt wird (z.B. andere Vorlage, gescannte/reine Bild-PDF)."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            if not pdf.pages:
                return None, None
            words = pdf.pages[0].extract_words()
    except Exception:
        return None, None

    beleg_label = next(
        (w for w in words if w["text"].startswith("Beleg-Nr")), None
    )
    if beleg_label is None:
        return None, None

    datum_label = next(
        (
            w
            for w in words
            if w["text"] == "Datum" and abs(w["top"] - beleg_label["top"]) < 2
        ),
        None,
    )

    beleg_nr = _wert_unter_label(words, beleg_label)
    datum = _wert_unter_label(words, datum_label) if datum_label else None
    return beleg_nr, datum


_ABHOLTAG_MUSTER = [
    re.compile(r"vom\s+(\d{2})\.(\d{2})\.(\d{4})"),
    re.compile(r"Abholtag\s+(\d{2})\.(\d{2})\.(\d{2,4})"),
]


def extrahiere_abholtag(pdf_bytes: bytes) -> Optional[str]:
    """Liest den Abholtag (TT.MM.JJJJ) aus dem Fliesstext (z.B. "Reise:
    ... vom 10.08.2026" oder "Abholtag 10.08.26"). Durchsucht alle
    Seiten, gibt None zurueck wenn kein passendes Muster gefunden wird."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                for muster in _ABHOLTAG_MUSTER:
                    match = muster.search(text)
                    if match:
                        tag, monat, jahr = match.groups()
                        if len(jahr) == 2:
                            jahr = "20" + jahr
                        return f"{tag}.{monat}.{jahr}"
    except Exception:
        return None
    return None


_BETRAG_MUSTER = re.compile(r"Endbetrag\s*\n?\s*([\d.,]+)\s*EUR")


def extrahiere_betrag(pdf_bytes: bytes) -> Optional[str]:
    """Liest den Endbetrag (z.B. "468,86 EUR") - steht meist auf der
    letzten Seite, daher werden alle Seiten durchsucht. Gibt None zurueck,
    wenn kein "Endbetrag" gefunden wird."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                match = _BETRAG_MUSTER.search(text)
                if match:
                    return f"{match.group(1)} EUR"
    except Exception:
        return None
    return None


_FIRMA_MUSTER = re.compile(r"Empf[aä]nger\s*:\s*(.+?)\s*\.\s*[A-Z]{1,2}-")


def extrahiere_firma(pdf_bytes: bytes) -> Optional[str]:
    """Liest die Empfaenger-Firma (z.B. "Skoda Auto a.s.") aus dem
    Fliesstext. Durchsucht alle Seiten, gibt None zurueck wenn nicht
    gefunden."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                match = _FIRMA_MUSTER.search(text)
                if match:
                    return match.group(1).strip()
    except Exception:
        return None
    return None


_REFERENZ_MUSTER = re.compile(r"Nummer:\s*\n?\s*([\d;]+)")


def extrahiere_referenznummern(pdf_bytes: bytes) -> List[str]:
    """Liest die SLB-Referenznummern (z.B. unter "SLB\\nNummer:") - eine
    durch Semikolon getrennte Liste von Nummern. Durchsucht alle Seiten,
    gibt eine leere Liste zurueck wenn nichts gefunden wird."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                match = _REFERENZ_MUSTER.search(text)
                if match:
                    return [n for n in match.group(1).split(";") if n]
    except Exception:
        return []
    return []


def archiv_monatsordner(archiv_basisordner_vorlage: str, abholtag: str) -> str:
    """Baut den Pfad zum Monatsordner (01-12) im Archiv fuer ein Datum
    (TT.MM.JJJJ). "{jahr}" in der Vorlage wird durch das Jahr des
    Abholtags ersetzt."""
    tag, monat, jahr = abholtag.split(".")
    basis = (
        archiv_basisordner_vorlage.format(jahr=jahr)
        if "{jahr}" in archiv_basisordner_vorlage
        else archiv_basisordner_vorlage
    )
    return os.path.join(basis, monat)


def _normalisiere(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", text.lower())


def erstelle_fallordner(
    archiv_basisordner_vorlage: str,
    reklamationen_basisordner: str,
    beleg_nr: str,
    abholtag: str,
    rechnung_pfad: str,
) -> Tuple[str, List[str]]:
    """Sucht die Planungs-/Ladelisten-/Avisierungs-Datei des Abholtags im
    passenden Monatsordner und kopiert sie zusammen mit der Rechnung in
    einen neuen Ordner "Fall-<Beleg-Nr>" unter reklamationen_basisordner.

    Gibt (fallordner_pfad, liste_nicht_gefundener_dateitypen) zurueck -
    fehlende Dateien fuehren nicht zum Abbruch, sondern werden nur
    gemeldet (Layout/Namenskonvention kann variieren)."""
    tag, monat, jahr = abholtag.split(".")
    datumspraefix = f"{jahr}{monat}{tag}"
    monatsordner = archiv_monatsordner(archiv_basisordner_vorlage, abholtag)

    fallordner = os.path.join(reklamationen_basisordner, f"{FALLORDNER_PRAEFIX}{beleg_nr}")
    os.makedirs(fallordner, exist_ok=True)
    shutil.copy2(rechnung_pfad, os.path.join(fallordner, os.path.basename(rechnung_pfad)))

    if os.path.isdir(monatsordner):
        vorhandene_dateien = os.listdir(monatsordner)
    else:
        vorhandene_dateien = []

    nicht_gefunden = []
    for typ, suchbegriff in PLANUNGSDATEI_TYPEN.items():
        treffer = next(
            (
                d
                for d in vorhandene_dateien
                if datumspraefix in _normalisiere(d)
                and _normalisiere(suchbegriff) in _normalisiere(d)
            ),
            None,
        )
        if treffer:
            shutil.copy2(
                os.path.join(monatsordner, treffer),
                os.path.join(fallordner, treffer),
            )
        else:
            nicht_gefunden.append(typ)

    return fallordner, nicht_gefunden


_UNSICHERE_ZEICHEN = re.compile(r'[\\/:*?"<>|]+')


def sicherer_pdf_dateiname(beleg_nr: Optional[str], fallback_dateiname: str) -> str:
    """Baut aus der Beleg-Nr. einen Dateinamen (Fallback: Original-Dateiname,
    falls die Beleg-Nr. nicht erkannt wurde)."""
    basis = beleg_nr if beleg_nr else os.path.splitext(fallback_dateiname)[0]
    basis = _UNSICHERE_ZEICHEN.sub("_", basis).strip("_") or "Beleg"
    return f"{basis}.pdf"


def _eindeutiger_dateiname(ziel_ordner: str, dateiname: str) -> str:
    """Haengt bei Namenskollision _1, _2, ... an, statt eine Datei zu ueberschreiben."""
    basis, endung = os.path.splitext(dateiname)
    kandidat = dateiname
    zaehler = 1
    while os.path.exists(os.path.join(ziel_ordner, kandidat)):
        kandidat = f"{basis}_{zaehler}{endung}"
        zaehler += 1
    return kandidat


def speichere_pdf(basisordner: str, dateiname: str, inhalt: bytes) -> str:
    """Speichert die PDF-Bytes im PDF-Unterordner und gibt den vollen Pfad zurueck."""
    ziel_ordner = pdf_ordner(basisordner)
    os.makedirs(ziel_ordner, exist_ok=True)
    eindeutiger_name = _eindeutiger_dateiname(ziel_ordner, dateiname)
    ziel_pfad = os.path.join(ziel_ordner, eindeutiger_name)
    with open(ziel_pfad, "wb") as f:
        f.write(inhalt)
    return ziel_pfad


def lade_excel(pfad: str) -> List[Dict[str, Any]]:
    """Liest bestehende Reklamationen aus der Excel-Datei (leer, falls keine existiert)."""
    if not pfad or not os.path.exists(pfad):
        return []

    wb = load_workbook(pfad)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    header = [cell.value for cell in ws[1]]

    rows: List[Dict[str, Any]] = []
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in raw_row):
            continue
        entry = dict(zip(header, raw_row))
        rows.append({col: entry.get(col) or "" for col in EXCEL_COLUMNS})
    return rows


def speichere_excel(pfad: str, rows: List[Dict[str, Any]]) -> None:
    """Schreibt alle Reklamationen als Excel-Datei (ueberschreibt bestehende Datei)."""
    ordner = os.path.dirname(pfad)
    if ordner:
        os.makedirs(ordner, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.append(EXCEL_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in EXCEL_COLUMNS])

    for col_idx, width in enumerate(EXCEL_SPALTENBREITEN, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    wb.save(pfad)
